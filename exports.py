"""
Funciones de exportacion (Excel/XLSX).
"""

import pandas as pd
import streamlit as st
from datetime import datetime as dt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


@st.cache_data(show_spinner=False)
def df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Reporte") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # --- Estilos ---
    header_fill = PatternFill("solid", fgColor="0F2A4A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D6DAE2"),
        right=Side(style="thin", color="D6DAE2"),
        top=Side(style="thin", color="D6DAE2"),
        bottom=Side(style="thin", color="D6DAE2"),
    )

    alt_fill_even = PatternFill("solid", fgColor="F6F7F9")
    alt_fill_odd = PatternFill("solid", fgColor="FFFFFF")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # --- Detectar tipo de columna para formato ---
    date_cols = set()
    currency_cols = set()
    for col_name in df.columns:
        name_lower = str(col_name).lower()
        if "fecha" in name_lower:
            date_cols.add(col_name)
        if any(k in name_lower for k in ["monto", "honorario", "importe"]):
            currency_cols.add(col_name)

    # --- Header ---
    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # --- Body ---
    for row_num, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        fill = alt_fill_even if row_num % 2 == 0 else alt_fill_odd
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = wrap_align

            # Formato moneda
            if col_name in currency_cols and cell.value:
                try:
                    val = str(cell.value).replace("$", "").replace(".", "").replace(",", ".").strip()
                    cell.value = float(val)
                    cell.number_format = '#,##0.00'
                except (ValueError, TypeError):
                    pass

            # Formato fecha
            if col_name in date_cols and cell.value:
                try:
                    for fmt in ["%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d"]:
                        try:
                            cell.value = dt.strptime(str(cell.value).strip(), fmt)
                            cell.number_format = 'DD/MM/YYYY'
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass

    # --- Freeze + filter ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Ajustes de ancho (max 60) ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for r in range(2, min(ws.max_row, 400) + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = min(60, max(10, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Print (landscape, fit) ---
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
